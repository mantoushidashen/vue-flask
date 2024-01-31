#!/usr/bin/python3
# -*- coding: utf-8 -*-

import time
import os
import xlrd
from openpyxl  import load_workbook, Workbook
class MyExcel:
    def __init__(self, analysis_file_name, sysrs_file_path, crs_file_path):
        # 客户需求文件
        self.crs_file = crs_file_path
        # 系统需求文件
        self.sysrs_file = sysrs_file_path
        # 分析之后返回给前端的文件名称
        self.analysis_file_name = analysis_file_name + '.xlsx'


# 软件Software      s_num
# 软件/硬件Software/ Hardware   sh_num
# 硬件Hardware  h_num
# 硬件/结构Hardware/Mechanical hm_num
# 结构Mechanical  m_num
# 软件/硬件/结构Software/Hardware/Mechanical shm_num

#返回的字典结构
# {
#     "data": crs_dic,
#     "s_num": xx,
#     "sh_num": xx,
#     "h_num": "xx",
#     "hm_num": "xx",
#     "m_num": "xx",
#     "shm_num": "xx"
# }

    def analysis_crs_excel(self):
        crs_access, crs_not_access, crs_condition_access = 0, 0, 0
        crs_draft, crs_review, crs_approve, crs_obsolete = 0, 0, 0, 0
        shinry_access, shinry_not_access, shinry_condition_access = 0, 0, 0
        crs_s_num, crs_sh_num, crs_h_num, crs_hm_num, crs_m_num, crs_shm_num = 0, 0, 0, 0, 0, 0
        
        WorkBook = xlrd.open_workbook(self.crs_file)
        sheet = WorkBook.sheet_by_name('Customer Requirements')
        crs_dic = {'data': {}}

        nrows = sheet.nrows
        for row_num in range(1, nrows):
            if sheet.cell_value(row_num, 2) != '标题\nHeading':
                crs_dic['data'][sheet.cell_value(row_num, 0)] = sheet.cell_value(row_num, 4)
                soft_type_value = sheet.cell_value(row_num, 6)
                # 计算需求领域各个选项的值
                if '软件Software' == soft_type_value:
                    crs_s_num += 1
                elif '软件/硬件Software/ Hardware' == soft_type_value:
                    crs_sh_num += 1
                elif '硬件Hardware' == soft_type_value:
                    crs_h_num += 1
                elif '硬件/结构Hardware/Mechanical' == soft_type_value:
                    crs_hm_num += 1
                elif '结构Mechanical' == soft_type_value:
                    crs_m_num += 1
                elif '软件/硬件/结构Software/Hardware/Mechanical' == soft_type_value:
                    crs_shm_num += 1

                # print(ws.cell(row=row, column=20).value)
                # 欣锐意见列: 接受Accep 不接受Not Accept  条件接受Accept with Condition
                disposal_value = sheet.cell_value(row_num, 19)
                if '接受Accept' == disposal_value:
                    shinry_access += 1
                elif '不接受Not Accept' == disposal_value:
                    shinry_not_access += 1
                elif '条件接受Accept with Condition' == disposal_value:
                    shinry_condition_access += 1

                # 客户意见列: 接受Accep 不接受Not Accept  条件接受Accept with Condition
                consume_comments_valuee = sheet.cell_value(row_num, 21)
                if '接受Accept' == consume_comments_valuee:
                    crs_access += 1
                elif '不接受Not Accept' == consume_comments_valuee:
                    crs_not_access += 1
                elif '条件接受Accept with Condition' == consume_comments_valuee:
                    crs_condition_access += 1

                # 客户需求列： 草稿Draft 评审中In review 已批准Approved 失效Obsolete
                consume_feature_value = sheet.cell_value(row_num, 23)
                if '草稿Draft' == consume_feature_value:
                    crs_draft += 1
                elif '评审中In review' == consume_feature_value:
                    crs_review += 1
                elif '已批准Approved' == consume_feature_value:
                    crs_approve += 1
                elif '失效Obsolete' == consume_feature_value:                                       
                    crs_obsolete += 1

        crs_dic["crs_total"] = len(crs_dic['data'])
        crs_dic["crs_access"], crs_dic["crs_not_access"], crs_dic["crs_condition_access"] = crs_access, crs_not_access, crs_condition_access
        crs_dic["crs_draft"], crs_dic["crs_review"], crs_dic["crs_approve"], crs_dic["crs_obsolete"] = crs_draft, crs_review, crs_approve, crs_obsolete
        crs_dic["shinry_access"], crs_dic["shinry_not_access"], crs_dic["shinry_condition_access"] = shinry_access, shinry_not_access, shinry_condition_access
        crs_dic["crs_s_num"], crs_dic["crs_sh_num"], crs_dic["crs_h_num"], crs_dic["crs_hm_num"], crs_dic["crs_m_num"], crs_dic["crs_shm_num"] = crs_s_num, crs_sh_num, crs_h_num, crs_hm_num, crs_m_num, crs_shm_num
        print(len(crs_dic['data']))
        return crs_dic

    def analysis_sysrs_excel(self):
        sysrs_draft, sysrs_review, sysrs_approve, sysrs_obsolete = 0, 0, 0, 0
        WorkBook = xlrd.open_workbook(self.sysrs_file)
        sheet = WorkBook.sheet_by_name('Requirements')
        sysrs_dic = {'data': {}}
        nrows = sheet.nrows

        for row_num in range(1, nrows):
            if sheet.cell_value(row_num, 4) != "标题Heading":
                status_value = sheet.cell_value(row_num, 23)
                if 'Draft草稿' == status_value:
                    sysrs_draft += 1
                elif 'In review评审中' == status_value:
                    sysrs_review += 1
                elif 'Approved已批准' == status_value:
                    sysrs_approve += 1
                elif 'Obsolete失效' == status_value:
                    sysrs_obsolete += 1
                sysrs_dic['data'][sheet.cell_value(row_num, 0)] = [[sheet.cell_value(row_num, 24), ], sheet.cell_value(row_num, 0)]


        sysrs_dic["sysrs_draft"], sysrs_dic["sysrs_review"], sysrs_dic["sysrs_approve"], sysrs_dic["sysrs_obsolete"] = sysrs_draft, sysrs_review, sysrs_approve, sysrs_obsolete
        return sysrs_dic



    def write_excl(self, data):
        file_dir = os.path.join(os.getcwd(), 'static')
        wb = Workbook()
        ws = wb.active
        row = 1
        
        for sysrs_value_list in dict(data).values():
            for sysrs_dic in sysrs_value_list:
                for crs_key, sysrs_value in sysrs_dic.items():
                    ws.cell(row=row, column=1).value = crs_key
                    ws.cell(row=row, column=2).value = sysrs_value[0]
                row += 1
        wb.save(os.path.join(file_dir, self.analysis_file_name))


    def merge_crs_sysrs(self):
        crs_data = self.analysis_crs_excel()
        sysrs_data = self.analysis_sysrs_excel()
        

        # 获取客户和需求需要被对比的数据
        crs_diff_data, sysrs_diff_data = crs_data['data'], sysrs_data['data']

        for sysrc_key, crs_value in sysrs_diff_data.items():
            if '\n' in crs_value[0][0]:
                crs_value_list = crs_value[0][0].split('\n')
                sysrs_diff_data[sysrc_key][0] = crs_value_list
        merge_dic = {}
        for key, value in sysrs_diff_data.items():
            crs_key_list = value[0]

            if len(crs_key_list) > 1:
                for crs_value in crs_key_list:
                    if merge_dic.get(crs_value):
                        merge_dic[crs_value].append({crs_value: value[1:]})
                    else:
                        merge_dic[crs_value] = [{crs_value: value[1:]}]
            else:
                if merge_dic.get(crs_key_list[0]):
                    merge_dic[crs_key_list[0]].append({crs_key_list[0]: value[1:]})
                else:
                    merge_dic[crs_key_list[0]] = [{crs_key_list[0]: value[1:]}]


        for crs_key in crs_diff_data.keys():
            if not merge_dic.get(crs_key):
                merge_dic[crs_key] = [{crs_key: ["\\"]}]
        self.write_excl(merge_dic)
        



        # 需要返回给前端画图的数据
        crs_del_data_dic = {key: value for key, value in crs_data.items() if key != 'data'}
        sysrs_del_data_dic = {key: value for key, value in sysrs_data.items() if key != 'data'}
        response_data = dict(crs_del_data_dic, **sysrs_del_data_dic)
        return response_data